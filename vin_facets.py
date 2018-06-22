import requests
import json
import openpyxl
from collections import Counter
import operator
from openpyxl import Workbook, load_workbook
from string_builders import *


def get_models(base_url, parameters):
    response = requests.get(base_url, params=parameters)
    data = response.json()
    facets = data["facets"]
    models_all = facets["model"]

    # Remove motorcycles:
    parameters["vehicle_type"] = "motorcycle"
    response = requests.get(base_url, params=parameters)
    data = response.json()
    facets = data["facets"]
    models_motorcycles = facets["model"]
    models = [x for x in models_all if x not in models_motorcycles]

    return models


def get_trims(base_url, parameters):
    response = requests.get(base_url, params=parameters)
    data = response.json()
    facets = data["facets"]
    trims = facets["trim"]

    if not trims:
        trims = [{"item": ""}]
    return trims


def get_alreadyCompleted(ws, year):
    completed_models = []
    for row in range(2, ws.max_row + 1):
        cell_name_year = "{}{}".format('A', row)
        if ws[cell_name_year].value == year:
            completed_models.append(ws["{}{}".format('C', row)].value)

    return list(set(completed_models))


base_url = "https://marketcheck-prod.apigee.net/v1/search"
# api_key = "aFHqYBK7moVJZN89peNCLNLKRnNN5Lxi" # patrick.cohen.at@gmail.com HOME-888-5
api_key = "fCevMoSHi1Uj5uJBcx5NG23IFqNZprV6" # kathryn.ann82@gmail.com XFinity
latitude = 38.783207
longitude = -95.920740
radius = 3000000

make = "Ford"

wb = load_workbook('/Users/patrick/Dropbox/byteRyde/02_Product/04_VIN/01_Results/'+make+'.xlsx')
ws = wb['Sheet1']

wb_stats = load_workbook("ApiStats.xlsx")
ws_stats = wb_stats.worksheets[0]

call_counter = ws_stats['A1'].value






years = range(1999, 2020)
# years = range(2002, 2015)
# years = [ 2015]
# years = [1999,2000,2001,2002,2003,2004,2005,2006,2007,2008,2009,2010,2011,2012,2013,2014,2015,2016,2017,2018,2019]
# years = [2013,2014,2015,2016,2017,2018,2019]
# years = [2016,2017,2018,2019]

for year in years:

    alreadyCompleted = get_alreadyCompleted(ws, year)
    # Get all models
    parameters = {"api_key": api_key,
                  "latitude": latitude,
                  "longitude": longitude,
                  "radius": radius,
                  "year": year,
                  "make": make,
                  "rows": 1,
                  "facets": "model"
                  }
    models = get_models(base_url, parameters)

    for model in models:
        model = model["item"]

        skipIt = 0
        alreadyDone = 0
        # if year == 2007 and make =="Cadillac" and model == "Escalade EXT":
        #     skipIt = 1
        #
        # if year == 2009 and make == "Cadillac" and model == "Escalade Hybrid":
        #     skipIt = 1
        #
        # if year == 2009 and make == "Cadillac" and model == "STS":
        #     skipIt = 1
        #
        # if year == 2017 and make == "Cadillac" and model == "ATS":
        #     skipIt = 1

        # if year == 2009 and make == "Hyundai" and model == "Azera":
        #     skipIt = 1

        # if year == 2014 and make == "Jeep" and model == "Compass":
        #     skipIt = 1
        #
        # if year == 2015 and make == "Jeep" and model == "Patriot":
        #     skipIt = 1
        #
        # if year == 2015 and make == "Jeep" and model == "Renegade":
        #     skipIt = 1
        #
        # if year == 2015 and make == "Jeep" and model == "Grand Cherokee SRT":
        #     skipIt = 1
        #
        # if year == 2016 and make == "Jeep" and model == "Cherokee":
        #     skipIt = 1

        # if year == 2000 and make == "Ford" and model == "F-250 Super Duty":
        #     skipIt = 1
        #
        # if year == 2000 and make == "Ford" and model == "Ranger":
        #     skipIt = 1
        #
        # if year == 2002 and make == "Ford" and model == "Escape":
        #     skipIt = 1
        #
        # if year == 2007 and make == "Ford" and model == "Ranger":
        #     skipIt = 1
        #

        # if model in alreadyCompleted:
        #     alreadyDone = 1
        #

        # if year == 2010 and make == "Ford":
        #     if model in ["F-150","Fusion", "Escape", "Edge", "Mustang", "Focus", "Explorer", "Taurus", "Expedition", "Ranger","F-250 SD", "Flex", "E-Series Van", "Fusion Hybrid", "Transit Connect", "F-350 Super Duty"]:
        #         alreadyDone = 1
        #
        # if year == 2011 and make == "Ford":
        #     if model in ["F-150","Edge", "Escape", "F-250 Super Duty", "Fusion", "Explorer", "Taurus", "Ranger",  "F-350 Super Duty", "Mustang", "Fiesta", "Expedition"]:
        #         alreadyDone = 1
        #
        # if year == 2012 and make == "Ford":
        #     if model in ["F-150","Focus", "Fusion", "Escape", "F-250 Super Duty", "Explorer", "Mustang", "Edge", "Taurus", "F-350 Super Duty"]:
        #         alreadyDone = 1
        #
        # if year == 2013 and make == "Ford":
        #     if model in ["F-150","Escape","Edge","Explorer","Fusion", "Focus", "Taurus", "Mustang","F-250 Super Duty", "Expedition", "Fiesta", "Flex","E-Series Van","F-350 Super Duty", "C-Max Hybrid"]:
        #         alreadyDone = 1
        #
        # if year == 2014 and make == "Ford":
        #     if model in ["F-150","Escape","Focus","Fusion","Edge","Explorer","Mustang","Flex","F-250 Super Duty","Expedition","Taurus","Fiesta","F-350 Super Duty","Transit Connect","Fusion Hybrid","E-Series Van","Focus ST","E-Series Wagon","C-Max Hybrid"]:
        #         alreadyDone = 1


        if skipIt == 1 or model in alreadyCompleted:
            print("skipping: " + str(year)
                  + " " + make + " " + model)
        else:

            print("num_calls: " + str(call_counter))
            print(str(year)+" "+make+" "+model)

            parameters = {"api_key": api_key,
                          "latitude": latitude,
                          "longitude": longitude,
                          "radius": radius,
                          "year": year,
                          "make": make,
                          "model": model,
                          "rows": 1,
                          "facets": "trim"
                          }

            trims = get_trims(base_url, parameters)


            for trim in trims:
                trim = trim["item"]

                parameters = {"api_key": api_key,
                              "latitude": latitude,
                              "longitude": longitude,
                              "radius": radius,
                              "year": year,
                              "make": make,
                              "model": model,
                              "trim": trim,
                              "rows": 1,
                              "facets": "trim,body_type,body_subtype,drivetrain,cylinders,transmission,doors,engine_type"
                              }

                response = requests.get(base_url, params=parameters)
                call_counter += 1
                if response.status_code != 200:
                    VinEntry_4Excel = [year, make, model,
                                       "Entire model",
                                       response.status_code, response.status_code, response.status_code]

                    ws.append(VinEntry_4Excel)

                else:
                    data = response.json()

                    facets = data["facets"]

                    # trims = facets["trim"]
                    # if not trims:
                    #     trims = [{"item": ""}]

                    body_types = facets["body_type"]
                    if not body_types:
                        body_types = [{"item": ""}]

                    body_subtypes = facets["body_subtype"]
                    if not body_subtypes:
                        body_subtypes = [{"item": ""}]

                    drivetrains = facets["drivetrain"]
                    if not drivetrains:
                        drivetrains = [{"item": ""}]

                    cylinders = facets["cylinders"]
                    if not cylinders:
                        cylinders = [{"item": ""}]
                    else:
                        # Remove entries that indicate a single cylinder (Audi issue)
                        singleCylIndex = next((index for (index, d) in enumerate(cylinders) if d["item"] == "1"), None)
                        if singleCylIndex:
                            del cylinders[singleCylIndex]

                        # Remove entries that indicate 15 cylinder (Honda issue)
                        FifteenCylIndex = next((index for (index, d) in enumerate(cylinders) if d["item"] == "15"), None)
                        if FifteenCylIndex:
                            del cylinders[FifteenCylIndex]


                    transmissions = facets["transmission"]
                    if not transmissions:
                        transmissions = [{"item": ""}]

                    doors = facets["doors"]
                    if not doors:
                        doors = [{"item": ""}]

                    engine_types = facets["engine_type"]
                    if not engine_types:
                        engine_types = [{"item": ""}]

                    # for trim in trims:
                    for body_type in body_types:
                        for body_subtype in body_subtypes:
                            for drivetrain in drivetrains:
                                for cylinder in cylinders:
                                    for transmission in transmissions:
                                        for door in doors:
                                            for engine_type in engine_types:

                                                parameters = {"api_key": api_key,
                                                              "latitude": latitude,
                                                              "longitude": longitude,
                                                              "radius": radius,
                                                              "year": year,
                                                              "make": make,
                                                              "model": model,
                                                              "rows": 30,
                                                              "trim": trim,
                                                              "body_type": body_type["item"],
                                                              "body_subtype": body_subtype["item"],
                                                              "drivetrain": drivetrain["item"],
                                                              "cylinder": cylinder["item"],
                                                              "transmission": transmission["item"],
                                                              "doors": door["item"],
                                                              "engine_type": engine_type["item"],
                                                              }
                                                response = requests.get(base_url, params=parameters)
                                                call_counter += 1
                                                if response.status_code ==200:
                                                    data = response.json()

                                                    VINs = []
                                                    if len(data["listings"]) > 0:
                                                        for listing in data["listings"]:
                                                            currVIN = listing["vin"][0:8]
                                                            VINs.append(currVIN)

                                                        VinBuckets = Counter(VINs)
                                                        VIN_use = max(VinBuckets.items(), key=operator.itemgetter(1))[0]
                                                        num_occurrence = VinBuckets[VIN_use]
                                                        num_all = len(VINs)

                                                        # ActualCylinderBuckets = Counter(int(float(data["listings"]["build"]["cylinders"])))
                                                        # ActualCylinders = data["listings"][0]["build"]["cylinders"]

                                                        VinEntry = {}
                                                        VinEntry["Year"] = year
                                                        VinEntry["Make"] = make
                                                        VinEntry["Model"] = model
                                                        VinEntry["bR_trim"] = buildTrimString(trims,body_types,body_subtypes,drivetrains,cylinders,transmissions,doors,engine_types,parameters)
                                                        VinEntry["VIN_1_3"] = VIN_use[0:3]
                                                        VinEntry["VIN_4_7"] = VIN_use[3:7]
                                                        VinEntry["VIN_8"] = VIN_use[7]

                                                        VinEntry_4Excel = [VinEntry["Year"], VinEntry["Make"], VinEntry["Model"],
                                                                           VinEntry["bR_trim"], VinEntry["VIN_1_3"], VinEntry["VIN_4_7"],
                                                                           VinEntry["VIN_8"], num_occurrence, num_all ]
                                                    else:
                                                        VinEntry_4Excel = [year, make, model, buildTrimString(trims, body_types, body_subtypes, drivetrains, cylinders, transmissions, doors, engine_types, parameters), "DNF", "DNF", "DNF"]
                                                else:
                                                    VinEntry_4Excel = [year, make, model, buildTrimString(trims, body_types, body_subtypes, drivetrains, cylinders, transmissions, doors, engine_types, parameters), response.status_code, response.status_code, response.status_code]

                                                ws.append(VinEntry_4Excel)
                                                ws_stats['A1'] = call_counter
                                                wb_stats.save("ApiStats.xlsx")

        wb.save('/Users/patrick/Dropbox/byteRyde/02_Product/04_VIN/01_Results/'+make+".xlsx")

